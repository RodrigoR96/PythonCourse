import openpyxl
from openpyxl import Workbook
import re
import sys
from whoisapi import *
import json
import os

path = os.path.dirname(os.path.realpath(__file__))
try:
    wsIp = openpyxl.load_workbook(path + "\\IPs.xlsx").active
except Exception as e:
    print("There is an issue with IPs.xlsx")
    print(e)
    sys.exit(1)
wsMaxRow = wsIp.max_row
ipList = []
validIpv4 = re.compile(r"^((25[0-5]|(2[0-4]|1[0-9]|[1-9]|)[0-9])(\.(?!$)|$)){4}$")
ipInvalidList = []
ipFailedList = []
wbRes = Workbook()
wsRes = wbRes.active
wic = Client(api_key='')

#// Read IPs from xlsx file.
for i in range(1, wsMaxRow + 1):
    ip = wsIp.cell(row = i, column = 1).value
    if not ip == "Servers" and not ip is None:
        #// Validate IPv4 with regex.
        if re.fullmatch(validIpv4, ip):
            ipList.append(ip)
        else:
            ipInvalidList.append(ip)
    else:
        ipInvalidList.append(ip)

#// TXT file creation
ipsTxt = open(path + "\\ips.txt", "w")
ipsTxt.write("ALL IPS\n")

for ip in ipList:#// Write all valid IPs into txt file
    ipsTxt.write(ip + "\n")

ipList = list(dict.fromkeys(ipList)) #// Remove duplicates by converting List into a Dictionary and then into a List again. Only duplicated: 52.217.96.244
ipsTxt.write("\n\n\nUNIQUE IPS\n")

for ip in ipList:#// Write all unique valid IPs into txt file
    ipsTxt.write(ip + "\n")

ipsTxt.close()

#// Whois-Api request and new excel and JSON file creation
wsRes.append(["path filename", "domainName", "registrarName", "contactEmail", "registryData.createdDate", "registrant.country"])
for ip in ipList:
    try:
        whois = wic.raw_data(ip)
        jsonWhois = json.loads(whois)
        #wsRes.append([jsonWhois["WhoisRecord"]["domainName"], jsonWhois["WhoisRecord"]["registrarName"], jsonWhois["WhoisRecord"]["contactEmail"], jsonWhois["WhoisRecord"]["registryData"]["createdDate"], jsonWhois["WhoisRecord"]["registryData"]["registrant"]["country"]])
        jsonFileLoc = path + "\\json\\" + ip + ".json"
        wsRes.append([ jsonFileLoc, jsonWhois.get("WhoisRecord", {}).get("domainName", " "), jsonWhois.get("WhoisRecord", {}).get("registrarName", " "), jsonWhois.get("WhoisRecord", {}).get("contactEmail", " "), jsonWhois.get("WhoisRecord", {}).get("registryData", {}).get("createdDate", " "), jsonWhois.get("WhoisRecord", {}).get("registryData", {}).get("registrant", {}).get("country", " ")])
        if not os.path.exists(path + "\\jsonFiles"):
            os.mkdir(path + "\\jsonFiles")
        with open(path + "\\jsonFiles\\"+ ip + ".json", "w") as f:   
            f.write(whois)
        
    except Exception as e:
        print(f"There were an error with: {ip}")
        print(str(e) + "\n")
        ipFailedList.append(ip)
        
wbRes.save(path + "\\serversInfo.xlsx")

#// Failed IPs TXT file creation
with open(path + "\\failed_ips.txt", "w") as f:
    for ip in ipFailedList:    
        f.write(ip + "\n")